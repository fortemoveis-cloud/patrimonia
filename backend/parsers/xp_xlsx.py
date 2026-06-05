import openpyxl
import io
import re
from datetime import date, datetime
from typing import List, Optional
from parsers.base import BaseParser, ParsedRecord


SECTION_MAP = {
    "termos": "equity",
    "aluguel": "equity",
    "coe": "fund",
    "pós-fixado": "fixed_income",
    "pos-fixado": "fixed_income",
    "prefixado": "fixed_income",
    "inflação": "fixed_income",
    "inflacao": "fixed_income",
    "compromissadas": "fixed_income",
    "fundos": "fund",
    "ações": "equity",
    "acoes": "equity",
    "renda variável": "equity",
    "renda variavel": "equity",
    "renda fixa": "fixed_income",
}


def _is_section_header(row_vals: list) -> Optional[str]:
    non_null = [c for c in row_vals if c is not None and str(c).strip()]
    if len(non_null) > 3:
        return None
    for val in non_null:
        s = str(val).strip().lower()
        for key in SECTION_MAP:
            if key in s:
                return SECTION_MAP[key]
    return None


def _br_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


class XpXlsxParser(BaseParser):
    def can_parse(self, filename: str, file_bytes: bytes) -> bool:
        name_lower = filename.lower()
        if not name_lower.endswith(".xlsx") and not name_lower.endswith(".xls"):
            return False
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            for name in wb.sheetnames:
                if "carteira" in name.lower() or "sua carteira" in name.lower():
                    ws = wb[name]
                    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                        row_text = " ".join(str(c) for c in row if c).lower()
                        if "patrimônio" in row_text or "patrimonio" in row_text:
                            return True
        except Exception:
            pass
        return False

    def _extract_date(self, filename: str, ws) -> date:
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, (date, datetime)):
                    return cell.date() if isinstance(cell, datetime) else cell
                m = re.search(r"(\d{2})/(\d{2})/(\d{4})", str(cell))
                if m:
                    try:
                        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                    except ValueError:
                        pass
        m = re.search(r"(\d{4})[-_]?(\d{2})[-_]?(\d{2})", filename)
        if m:
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass
        m = re.search(r"(\d{2})[-_](\d{2})[-_](\d{4})", filename)
        if m:
            try:
                return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                pass
        return date.today()

    def parse(self, filename: str, file_bytes: bytes) -> List[ParsedRecord]:
        # data_only=True (no read_only) so we can inspect cell.number_format
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        records: List[ParsedRecord] = []

        target_sheet = None
        for name in wb.sheetnames:
            if "carteira" in name.lower():
                target_sheet = name
                break
        if target_sheet is None:
            target_sheet = wb.sheetnames[0]

        ws = wb[target_sheet]
        snap_date = self._extract_date(filename, ws)

        # Load cell objects (for number_format) and extract values in parallel
        raw_rows = list(ws.iter_rows(values_only=False))
        rows = [tuple(cell.value for cell in row) for row in raw_rows]

        current_section = "fixed_income"
        is_aluguel = False
        aluguel_names: set = set()  # asset names found in "Aluguel" section (both sides net to zero)

        for row_idx, row in enumerate(rows):
            if all(c is None for c in row):
                continue

            non_null = [c for c in row if c is not None]
            if not non_null:
                continue

            row_vals = list(row)
            row_text = " ".join(str(c) for c in non_null).strip()

            # Detect section header
            section = _is_section_header(row_vals)
            if section is not None:
                current_section = section
                is_aluguel = "aluguel" in row_text.lower()
                continue

            # Skip summary/header rows
            lower_text = row_text.lower()
            if any(kw in lower_text for kw in ["total", "subtotal", "patrimônio", "patrimonio",
                                                "conta:", "produto", "ativo", "ticker"]):
                if "conta:" not in lower_text:
                    continue

            # First non-None cell = asset name
            first_val = None
            first_idx = 0
            for i, c in enumerate(row_vals):
                if c is not None and str(c).strip():
                    first_val = c
                    first_idx = i
                    break

            if first_val is None:
                continue

            asset_name = str(first_val).strip()
            if re.match(r"^[\d.,/\-]+$", asset_name):
                continue
            if len(asset_name) < 2:
                continue
            if re.match(r"^-?R\$\s*[\d.,]+$", asset_name):
                continue

            row_cells = raw_rows[row_idx] if row_idx < len(raw_rows) else ()

            # Collect monetary numerics — skip cells with % number_format (Excel pct fields)
            numeric_vals: List[float] = []
            date_vals: List[date] = []
            text_vals = [asset_name]

            for i, c in enumerate(row_vals[first_idx + 1:]):
                if c is None:
                    continue
                ci = first_idx + 1 + i
                cell = row_cells[ci] if ci < len(row_cells) else None
                nfmt = (getattr(cell, "number_format", None) or "") if cell else ""
                if "%" in nfmt:
                    continue  # skip Excel-formatted percentage cells

                d = _br_date(c)
                if d is not None:
                    date_vals.append(d)
                    continue
                cleaned = self.clean_float(c)
                if cleaned is not None:
                    numeric_vals.append(cleaned)
                elif str(c).strip():
                    text_vals.append(str(c).strip())

            if len(numeric_vals) < 2:
                continue

            maturity      = max(date_vals) if date_vals else None
            # Application date = earliest date when two dates present (aplicação + vencimento)
            purchase_dt   = min(date_vals) if len(date_vals) >= 2 and min(date_vals) != max(date_vals) else None
            rate_desc = None
            for tv in text_vals[1:]:
                if "%" in tv or "cdi" in tv.lower() or "ipca" in tv.lower() or "selic" in tv.lower():
                    rate_desc = tv
                    break

            if current_section == "equity":
                if is_aluguel:
                    # Aluguel (short) + long = net zero; exclude both sides entirely.
                    # The corresponding loan is tracked in the Empréstimos module.
                    aluguel_names.add(asset_name)
                    continue
                # Columns (after pct skip): [Posição, Última Cotação, Quantidade]
                market_value = numeric_vals[0] if len(numeric_vals) > 0 else None
                price = numeric_vals[1] if len(numeric_vals) > 1 else None
                units = numeric_vals[2] if len(numeric_vals) > 2 else None
                records.append(ParsedRecord(
                    institution_name="XP Investimentos",
                    institution_country="BR",
                    institution_currency="BRL",
                    asset_identifier=None,
                    asset_name=asset_name,
                    asset_type=current_section,
                    asset_currency="BRL",
                    snapshot_date=snap_date,
                    units=units,
                    price=price,
                    market_value=market_value,
                    raw_source_file=filename,
                ))
            elif current_section == "fund":
                # Columns (after pct skip): [Posição, Rendimento bruto, Valor aplicado]
                market_value = numeric_vals[0] if len(numeric_vals) > 0 else None
                unrealized = numeric_vals[1] if len(numeric_vals) > 1 else None
                cost_basis = numeric_vals[2] if len(numeric_vals) > 2 else market_value
                records.append(ParsedRecord(
                    institution_name="XP Investimentos",
                    institution_country="BR",
                    institution_currency="BRL",
                    asset_identifier=None,
                    asset_name=asset_name,
                    asset_type=current_section,
                    asset_currency="BRL",
                    snapshot_date=snap_date,
                    cost_basis=cost_basis,
                    market_value=market_value,
                    unrealized_gain=unrealized,
                    maturity_date=maturity,
                    rate_description=rate_desc,
                    raw_source_file=filename,
                ))
            else:
                # Columns (after pct skip): [Posição a mercado, Valor aplic., Valor aplic.orig,
                #                            Qtd, Preço Unit., IR, IOF]
                gross      = numeric_vals[0] if len(numeric_vals) > 0 else None
                cost_basis = numeric_vals[1] if len(numeric_vals) > 1 else gross
                ir_amount  = numeric_vals[5] if len(numeric_vals) > 5 else 0.0
                iof_amount = numeric_vals[6] if len(numeric_vals) > 6 else 0.0
                ir_total   = (ir_amount or 0.0) + (iof_amount or 0.0)
                tax_net_val = (
                    round(gross - ir_total, 2)
                    if gross is not None and ir_total > 0
                    else None
                )
                records.append(ParsedRecord(
                    institution_name="XP Investimentos",
                    institution_country="BR",
                    institution_currency="BRL",
                    asset_identifier=None,
                    asset_name=asset_name,
                    asset_type=current_section,
                    asset_currency="BRL",
                    snapshot_date=snap_date,
                    cost_basis=cost_basis,
                    market_value=gross,
                    tax_gross=gross,
                    tax_net=tax_net_val,
                    maturity_date=maturity,
                    purchase_date=purchase_dt,
                    rate_description=rate_desc,
                    raw_source_file=filename,
                ))

        # Aggregate multi-lot positions with same name+type
        seen: dict = {}
        final: List[ParsedRecord] = []
        for r in records:
            key = (r.asset_name, r.asset_type)
            if key in seen:
                old = final[seen[key]]
                mv = (old.market_value or 0.0) + (r.market_value or 0.0)
                cb = ((old.cost_basis or 0.0) + (r.cost_basis or 0.0)
                      if (old.cost_basis is not None or r.cost_basis is not None) else None)
                u = ((old.units or 0.0) + (r.units or 0.0)
                     if (old.units is not None or r.units is not None) else None)
                # tax_net: sum net values (fall back to market_value per lot when no IR)
                if old.tax_net is not None or r.tax_net is not None:
                    tn = (old.tax_net if old.tax_net is not None else (old.market_value or 0.0)) \
                       + (r.tax_net   if r.tax_net   is not None else (r.market_value   or 0.0))
                else:
                    tn = None
                # purchase_date: earliest application date across lots
                dates_pd = [d for d in [old.purchase_date, r.purchase_date] if d is not None]
                pd_combined = min(dates_pd) if dates_pd else None
                final[seen[key]] = ParsedRecord(
                    institution_name=old.institution_name,
                    institution_country=old.institution_country,
                    institution_currency=old.institution_currency,
                    asset_identifier=old.asset_identifier,
                    asset_name=old.asset_name,
                    asset_type=old.asset_type,
                    asset_currency=old.asset_currency,
                    snapshot_date=old.snapshot_date,
                    units=u,
                    price=old.price,
                    cost_basis=cb,
                    market_value=mv,
                    tax_gross=mv,
                    tax_net=tn,
                    unrealized_gain=old.unrealized_gain,
                    maturity_date=old.maturity_date,
                    purchase_date=pd_combined,
                    rate_description=old.rate_description,
                    raw_source_file=old.raw_source_file,
                )
            else:
                seen[key] = len(final)
                final.append(r)

        # Exclude the long side of any aluguel-covered position (net is zero).
        return [
            r for r in final
            if not (r.asset_type == "equity" and r.asset_name in aluguel_names)
        ]
