import openpyxl
import io
import re
from datetime import date, datetime
from typing import List, Optional
from parsers.base import BaseParser, ParsedRecord

# Account number that distinguishes this parser from XP Empresa
ACCOUNT_NUMBER = "6200983"
INSTITUTION_NAME = "XP Investimentos (Pessoal)"

# Section type map — includes "fii" for Fundos Imobiliários (separate column layout)
_SECTION_MAP = {
    "fundos imobiliários": "fii",
    "fundos imobiliarios": "fii",
    "fundos de investimento": "fund",
    "fundos":               "fund",
    "coe":                  "fund",
    "pós-fixado":           "fixed_income",
    "pos-fixado":           "fixed_income",
    "prefixado":            "fixed_income",
    "inflação":             "fixed_income",
    "inflacao":             "fixed_income",
    "renda fixa":           "fixed_income",
    "ações":                "equity",
    "acoes":                "equity",
    "renda variável":       "equity",
    "renda variavel":       "equity",
    "termos":               "equity",
}

# Keywords that identify column-header or summary rows to skip
_SKIP_KEYWORDS = [
    "total", "subtotal", "patrimônio", "patrimonio",
    "produto", "ativo", "ticker",
    "posição na taxa",      # Renda Fixa column header first cell variant
    "data aplicação",       # column header row keyword
    "quantidade de cotas",  # FII column header row keyword
    "taxa de compra",       # column header row keyword
    "conta:",
]


def _is_section_header(row_vals: list) -> Optional[str]:
    non_null = [c for c in row_vals if c is not None and str(c).strip()]
    if len(non_null) > 3:
        return None
    for val in non_null:
        s = str(val).strip().lower()
        for key, section in _SECTION_MAP.items():
            if key in s:
                return section
    return None


def _br_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


class XpPersonalXlsxParser(BaseParser):
    """
    Parser for XP Investimentos personal account (conta 6200983).
    Handles three section layouts found in PosicaoDetalhada files:

    Renda Fixa (Pós-Fixado):
      cols (after % skip): [gross, cost_basis, orig_cost_basis, qty, price, IR, IOF, tax_net]
      dates: [Data aplicação (purchase), Data vencimento (maturity)]

    Fundos de Investimento:
      cols (after % skip): [gross, cost_basis, tax_net]
      (Rentabilidade Líquida/Bruta are % formatted → skipped)

    Fundos Imobiliários:
      cols (after % skip): [posição, preço_médio, última_cotação, quantidade]
      cost_basis = preço_médio × quantidade
    """

    def can_parse(self, filename: str, file_bytes: bytes) -> bool:
        if not filename.lower().endswith((".xlsx", ".xls")):
            return False
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
                    row_text = " ".join(str(c) for c in row if c)
                    if ACCOUNT_NUMBER in row_text:
                        return True
        except Exception:
            pass
        return False

    def _find_main_sheet(self, wb) -> str:
        """Return the sheet name that contains the account number, else first sheet."""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
                row_text = " ".join(str(c) for c in row if c)
                if ACCOUNT_NUMBER in row_text:
                    return sheet_name
        return wb.sheetnames[0]

    def _extract_date(self, filename: str, ws) -> date:
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, (date, datetime)):
                    return cell.date() if isinstance(cell, datetime) else cell
                m = re.search(r"(\d{2})/(\d{2})/(\d{4})", str(cell))
                if m:
                    try:
                        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                    except ValueError:
                        pass
        m = re.search(r"(\d{4})[-_]?(\d{2})[-_]?(\d{2})", filename)
        if m:
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass
        return date.today()

    def parse(self, filename: str, file_bytes: bytes) -> List[ParsedRecord]:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        records: List[ParsedRecord] = []

        target_sheet = self._find_main_sheet(wb)
        ws = wb[target_sheet]
        snap_date = self._extract_date(filename, ws)

        raw_rows = list(ws.iter_rows(values_only=False))
        rows = [tuple(cell.value for cell in row) for row in raw_rows]

        current_section = "fixed_income"

        for row_idx, row in enumerate(rows):
            if all(c is None for c in row):
                continue
            non_null = [c for c in row if c is not None]
            if not non_null:
                continue

            row_vals = list(row)
            row_text = " ".join(str(c) for c in non_null).strip()
            lower_text = row_text.lower()

            # Detect and set section type
            section = _is_section_header(row_vals)
            if section is not None:
                current_section = section
                continue

            # Skip header / summary rows
            if any(kw in lower_text for kw in _SKIP_KEYWORDS):
                continue

            # First non-None cell = asset name
            first_val = None
            first_idx = 0
            for i, c in enumerate(row_vals):
                if c is not None and str(c).strip():
                    first_val = c
                    first_idx = i
                    break

            if first_val is None:
                continue

            asset_name = str(first_val).strip()
            if re.match(r"^[\d.,/\-]+$", asset_name):
                continue
            if len(asset_name) < 2:
                continue
            if re.match(r"^-?R\$\s*[\d.,]+$", asset_name):
                continue

            row_cells = raw_rows[row_idx] if row_idx < len(raw_rows) else ()

            # Collect numeric and date values; skip Excel %-formatted cells
            numeric_vals: List[float] = []
            date_vals: List[date] = []
            text_vals = [asset_name]

            for i, c in enumerate(row_vals[first_idx + 1:]):
                if c is None:
                    continue
                ci = first_idx + 1 + i
                cell = row_cells[ci] if ci < len(row_cells) else None
                nfmt = (getattr(cell, "number_format", None) or "") if cell else ""
                if "%" in nfmt:
                    continue

                d = _br_date(c)
                if d is not None:
                    date_vals.append(d)
                    continue
                cleaned = self.clean_float(c)
                if cleaned is not None:
                    numeric_vals.append(cleaned)
                elif str(c).strip():
                    text_vals.append(str(c).strip())

            if len(numeric_vals) < 1:
                continue

            # Rate description (from text cells containing CDI/IPCA/SELIC/%)
            rate_desc: Optional[str] = None
            for tv in text_vals[1:]:
                if "%" in tv or "cdi" in tv.lower() or "ipca" in tv.lower() or "selic" in tv.lower():
                    rate_desc = tv
                    break

            if current_section == "fii":
                self._append_fii(records, asset_name, numeric_vals, snap_date, filename)

            elif current_section == "fund":
                self._append_fund(records, asset_name, numeric_vals, snap_date, filename)

            elif current_section == "equity":
                self._append_equity(records, asset_name, numeric_vals, snap_date, filename)

            else:
                # fixed_income (default)
                maturity    = max(date_vals) if date_vals else None
                purchase_dt = (min(date_vals)
                               if len(date_vals) >= 2 and min(date_vals) != max(date_vals)
                               else None)
                self._append_fixed(records, asset_name, numeric_vals,
                                   snap_date, filename, maturity, purchase_dt, rate_desc)

        return self._merge_lots(records)

    # ── Section-specific append helpers ────────────────────────────────────────

    def _append_fixed(self, records, name, nv, snap_date, filename,
                      maturity, purchase_dt, rate_desc):
        """
        Renda Fixa columns (after % skip):
          [0] Posição na taxa de compra (gross / tax_gross)
          [1] Valor aplicado (cost_basis)
          [2] Valor aplicado original
          [3] Quantidade
          [4] Preço Unitário
          [5] IR
          [6] IOF
          [7] Valor Líquido (explicit tax_net — preferred)
        """
        gross      = nv[0] if len(nv) > 0 else None
        cost_basis = nv[1] if len(nv) > 1 else gross
        ir_amount  = nv[5] if len(nv) > 5 else 0.0
        iof_amount = nv[6] if len(nv) > 6 else 0.0
        ir_total   = (ir_amount or 0.0) + (iof_amount or 0.0)

        # Use explicit Valor Líquido when available (column count = 8)
        if len(nv) > 7:
            tax_net = nv[7]
        elif gross is not None and ir_total > 0:
            tax_net = round(gross - ir_total, 2)
        else:
            tax_net = None

        records.append(ParsedRecord(
            institution_name=INSTITUTION_NAME,
            institution_country="BR",
            institution_currency="BRL",
            asset_identifier=None,
            asset_name=name,
            asset_type="fixed_income",
            asset_currency="BRL",
            snapshot_date=snap_date,
            cost_basis=cost_basis,
            market_value=gross,
            tax_gross=gross,
            tax_net=tax_net,
            maturity_date=maturity,
            purchase_date=purchase_dt,
            rate_description=rate_desc,
            raw_source_file=filename,
        ))

    def _append_fund(self, records, name, nv, snap_date, filename):
        """
        Fundos de Investimento columns (after % skip):
          [0] Posição (gross)
          [1] Valor aplicado (cost_basis)
          [2] Valor líquido (tax_net)
        Rentabilidade Líquida/Bruta are % formatted → skipped.
        If they are NOT skipped, fall back to last-value heuristic.
        """
        gross = nv[0] if len(nv) > 0 else None
        if len(nv) >= 3:
            # Most common: [posição, valor_aplicado, valor_liquido]
            cost_basis = nv[-2]
            tax_net    = nv[-1]
        elif len(nv) == 2:
            cost_basis = nv[1]
            tax_net    = nv[1]
        else:
            cost_basis = gross
            tax_net    = gross

        unrealized = round(gross - cost_basis, 2) if (gross and cost_basis) else None

        records.append(ParsedRecord(
            institution_name=INSTITUTION_NAME,
            institution_country="BR",
            institution_currency="BRL",
            asset_identifier=None,
            asset_name=name,
            asset_type="fund",
            asset_currency="BRL",
            snapshot_date=snap_date,
            cost_basis=cost_basis,
            market_value=gross,
            tax_gross=gross,
            tax_net=tax_net,
            unrealized_gain=unrealized,
            raw_source_file=filename,
        ))

    def _append_fii(self, records, name, nv, snap_date, filename):
        """
        Fundos Imobiliários columns (after % skip):
          [0] Posição (market_value = last_price × units)
          [1] Preço médio (avg purchase price)
          [2] Última cotação (last price)
          [3] Quantidade de Cotas (units)
        Rentabilidade c/ proventos and Bruta are % formatted → skipped.
        """
        if len(nv) < 2:
            return
        market_value = nv[0]
        # Use last-position heuristic in case rentabilidade cols are NOT % formatted:
        # last value = units, second-to-last = last_price, third-to-last = avg_price
        units      = nv[-1] if len(nv) >= 1 else None
        last_price = nv[-2] if len(nv) >= 2 else None
        avg_price  = nv[-3] if len(nv) >= 3 else None

        cost_basis = round(avg_price * units, 2) if (avg_price and units) else None
        unrealized = round(market_value - cost_basis, 2) if (market_value and cost_basis) else None

        records.append(ParsedRecord(
            institution_name=INSTITUTION_NAME,
            institution_country="BR",
            institution_currency="BRL",
            asset_identifier=None,
            asset_name=name,
            asset_type="equity",
            asset_currency="BRL",
            snapshot_date=snap_date,
            units=units,
            price=last_price,
            cost_basis=cost_basis,
            market_value=market_value,
            unrealized_gain=unrealized,
            raw_source_file=filename,
        ))

    def _append_equity(self, records, name, nv, snap_date, filename):
        """
        Ações / Renda Variável columns (after % skip):
          [0] Posição, [1] Última Cotação, [2] Quantidade
        """
        market_value = nv[0] if len(nv) > 0 else None
        price        = nv[1] if len(nv) > 1 else None
        units        = nv[2] if len(nv) > 2 else None

        records.append(ParsedRecord(
            institution_name=INSTITUTION_NAME,
            institution_country="BR",
            institution_currency="BRL",
            asset_identifier=None,
            asset_name=name,
            asset_type="equity",
            asset_currency="BRL",
            snapshot_date=snap_date,
            units=units,
            price=price,
            market_value=market_value,
            raw_source_file=filename,
        ))

    # ── Multi-lot aggregation ───────────────────────────────────────────────────

    def _merge_lots(self, records: List[ParsedRecord]) -> List[ParsedRecord]:
        """Combine multiple lots of the same asset into a single record."""
        seen: dict = {}
        final: List[ParsedRecord] = []

        for r in records:
            key = (r.asset_name, r.asset_type)
            if key not in seen:
                seen[key] = len(final)
                final.append(r)
                continue

            old = final[seen[key]]
            mv  = (old.market_value or 0.0) + (r.market_value or 0.0)
            cb  = ((old.cost_basis  or 0.0) + (r.cost_basis  or 0.0)
                   if (old.cost_basis  is not None or r.cost_basis  is not None) else None)
            u   = ((old.units  or 0.0) + (r.units  or 0.0)
                   if (old.units  is not None or r.units  is not None) else None)

            # tax_net: sum net values
            if old.tax_net is not None or r.tax_net is not None:
                tn = ((old.tax_net if old.tax_net is not None else (old.market_value or 0.0))
                    + (r.tax_net   if r.tax_net   is not None else (r.market_value   or 0.0)))
            else:
                tn = None

            # purchase_date: earliest across lots
            pds = [d for d in [old.purchase_date, r.purchase_date] if d is not None]
            pd  = min(pds) if pds else None

            final[seen[key]] = ParsedRecord(
                institution_name=old.institution_name,
                institution_country=old.institution_country,
                institution_currency=old.institution_currency,
                asset_identifier=old.asset_identifier,
                asset_name=old.asset_name,
                asset_type=old.asset_type,
                asset_currency=old.asset_currency,
                snapshot_date=old.snapshot_date,
                units=u,
                price=old.price,
                cost_basis=cb,
                market_value=mv,
                tax_gross=mv,
                tax_net=tn,
                unrealized_gain=old.unrealized_gain,
                maturity_date=old.maturity_date,
                purchase_date=pd,
                rate_description=old.rate_description,
                raw_source_file=old.raw_source_file,
            )

        return final
