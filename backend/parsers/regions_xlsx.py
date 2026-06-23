import openpyxl
import io
import re
from datetime import date
from typing import List, Optional
from parsers.base import BaseParser, ParsedRecord

_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_MATURITY_RE = re.compile(
    r"(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})\s*$",
    re.IGNORECASE,
)


def _parse_maturity(asset_name: str) -> Optional[date]:
    m = _MATURITY_RE.search(asset_name)
    if not m:
        return None
    try:
        day = int(m.group(1))
        month = _MONTH_MAP[m.group(2).lower()]
        year = int(m.group(3))
        return date(year, month, day)
    except (ValueError, KeyError):
        return None


CATEGORY_MAP = {
    "Fixed Income": "fixed_income",
    "Equity": "equity",
    "Cash and Equivalents": "cash",
    "Fund": "fund",
    "Alternative Investments": "fund",
}


class RegionsXlsxParser(BaseParser):
    def can_parse(self, filename: str, file_bytes: bytes) -> bool:
        if not filename.lower().endswith(".xlsx"):
            return False
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            for name in wb.sheetnames:
                if name.startswith("Investment_temp_"):
                    return True
            # Check cell content for ORQUIDEA marker
            for name in wb.sheetnames:
                ws = wb[name]
                for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                    for cell in row:
                        if cell and "ORQUIDEA INVESTMENTS" in str(cell).upper():
                            return True
        except Exception:
            pass
        return False

    def _extract_date_from_content(self, wb) -> Optional[date]:
        """Try to find the portfolio date inside the xlsx (sheet name or header cells)."""
        from datetime import datetime
        # Sheet name suffix: Investment_temp_20260610 or Investment_temp_06/10/2026
        for name in wb.sheetnames:
            if name.startswith("Investment_temp_"):
                suffix = name[len("Investment_temp_"):]
                m = re.match(r"(\d{4})(\d{2})(\d{2})$", suffix)
                if m:
                    try:
                        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    except ValueError:
                        pass
                m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", suffix)
                if m:
                    try:
                        return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
                    except ValueError:
                        pass
        # First rows of target sheet: look for "as of MM/DD/YYYY" or "as of Month DD, YYYY"
        target = None
        for name in wb.sheetnames:
            if name.startswith("Investment_temp_"):
                target = wb[name]
                break
        if target is None and wb.sheetnames:
            target = wb[wb.sheetnames[0]]
        if target:
            for row in target.iter_rows(min_row=1, max_row=15, values_only=True):
                for cell in row:
                    if not cell:
                        continue
                    s = str(cell).strip()
                    m = re.search(r"as\s+of\s+(\d{1,2})/(\d{1,2})/(\d{4})", s, re.IGNORECASE)
                    if m:
                        try:
                            return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
                        except ValueError:
                            pass
                    m = re.search(r"as\s+of\s+(\w+)\s+(\d{1,2}),?\s*(\d{4})", s, re.IGNORECASE)
                    if m:
                        try:
                            return datetime.strptime(
                                f"{m.group(1)} {m.group(2)}, {m.group(3)}", "%B %d, %Y"
                            ).date()
                        except ValueError:
                            pass
        return None

    def _extract_date_from_filename(self, filename: str) -> Optional[date]:
        # Regions portal: Investments_MM-DD-YYYY_HH-MM-SS_{am/pm}.xlsx
        # Anchor by underscores — prevents matching time component (e.g. 12-04-09 after the year)
        m = re.search(r"_(\d{2})-(\d{2})-(\d{4})_", filename)
        if m:
            try:
                return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
            except ValueError:
                pass
        # ISO with word boundaries (YYYY-MM-DD not inside a longer digit sequence)
        m = re.search(r"(?<!\d)(\d{4})-(\d{2})-(\d{2})(?!\d)", filename)
        if m:
            try:
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass
        return None

    def parse(self, filename: str, file_bytes: bytes) -> List[ParsedRecord]:
        # read_only=True truncates rows/cols for files without default styles; use default mode
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        snap_date = self._extract_date_from_content(wb) or self._extract_date_from_filename(filename) or date.today()
        records: List[ParsedRecord] = []

        target_sheet = None
        for name in wb.sheetnames:
            if name.startswith("Investment_temp_"):
                target_sheet = name
                break
        if target_sheet is None:
            target_sheet = wb.sheetnames[0]

        ws = wb[target_sheet]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row
        header_row_idx = None
        headers = {}
        for i, row in enumerate(rows):
            # Normalize: strip and collapse internal whitespace
            row_str = [" ".join(str(c).split()) if c is not None else "" for c in row]
            if "Investment Asset Name" in row_str or "Asset Name" in row_str:
                header_row_idx = i
                for j, h in enumerate(row_str):
                    headers[h] = j
                break

        if header_row_idx is None:
            return records

        col = lambda name: headers.get(name)

        for row in rows[header_row_idx + 1:]:
            def v(name):
                idx = col(name)
                if idx is None:
                    return None
                val = row[idx] if idx < len(row) else None
                return val

            asset_name = str(v("Investment Asset Name") or v("Asset Name") or "").strip()
            if not asset_name or asset_name.lower() in ("none", "nan", ""):
                continue

            category_raw = str(v("Investment Category") or "").strip()
            asset_type = CATEGORY_MAP.get(category_raw, "other")

            # Skip the loan-annotation row (not an investable asset)
            if "PLEDGED TO SECURE" in asset_name.upper():
                continue

            identifier = str(v("Primary Asset Identifier") or "").strip()
            portfolio_name = str(v("Portfolio Name") or "").strip()
            currency = str(v("Currency") or "USD").strip()

            records.append(ParsedRecord(
                institution_name="Regions Bank",
                institution_country="US",
                institution_currency="USD",
                asset_identifier=identifier or None,
                asset_name=asset_name,
                asset_type=asset_type,
                asset_currency=currency,
                snapshot_date=snap_date,
                units=self.clean_float(v("Units")),
                price=self.clean_float(v("Price")),
                cost_basis=self.clean_float(v("Cost Basis")),
                market_value=self.clean_float(v("Market Value")),
                unrealized_gain=self.clean_float(v("Unrealized gain/loss amount")),
                estimated_income=self.clean_float(v("Estimated Annual Income")),
                accrued_income=self.clean_float(v("Accrued Income")),
                current_yield=self.clean_float(v("Current Yield")),
                maturity_date=_parse_maturity(asset_name),
                portfolio_name=portfolio_name or None,
                raw_source_file=filename,
            ))

        return records
